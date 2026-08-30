"""
report_utils.py
----------------
Shared "business-friendly" Excel report builder used by both the
classification comparison (cli.py) and the data-extraction comparison
(extraction/cli.py). Callers build their own metric rows; this module
only owns the workbook layout/styling so it isn't duplicated per task.
"""

SECTION = "__SECTION__"


def fmt(val, kind=None):
    """Format a metric value; None -> 'N/A'."""
    if val is None:
        return "N/A"
    if kind == "pct":
        return f"{val * 100:.1f}%"
    if kind == "ms":
        return f"{val:.1f} ms"
    if kind == "usd":
        return f"${val:.4f}"
    return val


def fmt_rate(count, total):
    """Return 'XX.X%' or 'N/A' for count / total."""
    if count is None or total is None or total == 0:
        return "N/A"
    return f"{count / total * 100:.1f}%"


def write_comparison_excel(rows, out_path, sheet_title, col_a_header, col_b_header):
    """
    Writes a two-column (e.g. BERT vs LLM) business comparison workbook.

    `rows` is a list of (row_type, metric, business_description, value_a, value_b)
    tuples. A row with row_type == SECTION is rendered as a full-width section
    banner (metric holds the section title, the rest are ignored).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  [warn] openpyxl not installed; skipping Excel export. Run: pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    sec_font = Font(bold=True, color="FFFFFF", size=10)
    body_font = Font(size=10)

    hdr_fill = PatternFill("solid", fgColor="1F4E79")    # dark navy
    sec_fill = PatternFill("solid", fgColor="2E75B6")    # mid blue
    alt_fill = PatternFill("solid", fgColor="F5F5F5")    # very light gray
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="top", wrap_text=True)

    ws.append(["Metric", "Business Description", col_a_header, col_b_header])
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border

    ws.cell(row=1, column=3).fill = PatternFill("solid", fgColor="155A8A")
    ws.cell(row=1, column=4).fill = PatternFill("solid", fgColor="B8860B")

    ws.freeze_panes = "A2"

    data_row = 2
    alt = False
    for row in rows:
        row_type, metric, desc, val_a, val_b = row
        if row_type == SECTION:
            ws.append([metric, "", "", ""])
            sec_row = data_row
            for col in range(1, 5):
                cell = ws.cell(row=sec_row, column=col)
                cell.font = sec_font
                cell.fill = sec_fill
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border
            ws.merge_cells(f"A{sec_row}:D{sec_row}")
            data_row += 1
            alt = False
            continue

        ws.append([metric, desc, val_a, val_b])
        row_fill = alt_fill if alt else white_fill
        for col in range(1, 5):
            cell = ws.cell(row=data_row, column=col)
            cell.font = body_font
            cell.fill = row_fill
            cell.alignment = wrap
            cell.border = border
        ws.cell(row=data_row, column=1).font = Font(bold=True, size=10)
        ws.cell(row=data_row, column=3).fill = PatternFill("solid", fgColor="EBF3FB" if not alt else "D6E4F7")
        ws.cell(row=data_row, column=4).fill = PatternFill("solid", fgColor="FFFDE7" if not alt else "FFF9C4")
        data_row += 1
        alt = not alt

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 32

    for i in range(2, data_row):
        ws.row_dimensions[i].height = 40

    wb.save(out_path)
    print(f"  Excel report written to {out_path}")
